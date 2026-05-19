#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel题库转JavaScript数据文件
读取两个xlsx，输出 questions.js (BANKS 格式)
"""

import json
import re
from openpyxl import load_workbook

BASE_DIR = r"F:/XJ/claude/工程师题库"
OUTPUT_PATH = f"{BASE_DIR}/questions.js"


def parse_options_pipe(question_type, options_text):
    """解析旧题库选项（按 | 分割）"""
    if question_type == "判断题":
        return [{"key": "正确", "text": "正确"}, {"key": "错误", "text": "错误"}]
    parts = [p.strip() for p in str(options_text or "").split("|") if p.strip()]
    options = []
    for part in parts:
        match = re.match(r'^([A-Z])[\.．、]\s*(.+)$', part)
        if match:
            options.append({"key": match.group(1), "text": match.group(2).strip()})
        else:
            options.append({"key": part[0] if part else "", "text": part})
    return options


def parse_options_newline(question_type, options_text):
    """解析新题库选项（按 \n 分割）"""
    if question_type == "判断题":
        return [{"key": "正确", "text": "正确"}, {"key": "错误", "text": "错误"}]
    parts = [p.strip() for p in str(options_text or "").split("\n") if p.strip()]
    options = []
    for part in parts:
        match = re.match(r'^([A-Z])[\.．、]\s*(.+)$', part)
        if match:
            options.append({"key": match.group(1), "text": match.group(2).strip()})
        else:
            options.append({"key": part[0] if part else "", "text": part})
    return options


def parse_bank_old():
    path = f"{BASE_DIR}/装饰工程网上学习题库.xlsx"
    wb = load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    questions = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if len(row) < 5:
            continue
        seq, q_type, title, options_text, answer = row[:5]
        if not title or not q_type:
            continue
        questions.append({
            "id": int(seq) if isinstance(seq, (int, float)) else len(questions) + 1,
            "type": str(q_type).strip(),
            "title": str(title).strip(),
            "options": parse_options_pipe(str(q_type).strip(), options_text),
            "answer": str(answer).strip() if answer else ""
        })
    return {
        "id": "mid-engineer",
        "name": "中级工程师",
        "hasChapters": False,
        "questions": questions
    }


def parse_bank_new():
    path = f"{BASE_DIR}/课后习题汇总.xlsx"
    wb = load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    questions = []
    chapter_order = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 6:
            continue
        chapter, lesson, q_type, title, options_text, answer = row[:6]
        if not title or not q_type:
            continue
        chapter = str(chapter).strip()
        if chapter and chapter not in chapter_order:
            chapter_order.append(chapter)
        questions.append({
            "id": len(questions) + 1,
            "type": str(q_type).strip(),
            "title": str(title).strip(),
            "options": parse_options_newline(str(q_type).strip(), options_text),
            "answer": str(answer).strip() if answer else "",
            "chapter": chapter,
            "lesson": str(lesson).strip() if lesson else ""
        })
    return {
        "id": "erjian-edu",
        "name": "二建继续教育",
        "hasChapters": True,
        "chapters": chapter_order,
        "questions": questions
    }


def main():
    old_bank = parse_bank_old()
    new_bank = parse_bank_new()
    banks = [old_bank, new_bank]

    js_content = "const BANKS = " + json.dumps(banks, ensure_ascii=False, indent=2) + ";\n"
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        f.write(js_content)

    for bank in banks:
        print(f"题库: {bank['name']} ({bank['id']})")
        if bank.get("hasChapters"):
            print(f"  章节: {len(bank['chapters'])}")
        type_counts = {}
        for q in bank["questions"]:
            type_counts[q["type"]] = type_counts.get(q["type"], 0) + 1
        for t, c in type_counts.items():
            print(f"  {t}: {c} 题")
        print(f"  总计: {len(bank['questions'])} 题")
    print(f"输出: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
